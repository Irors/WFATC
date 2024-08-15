from data.constant import PATH
import openpyxl


def get_bd():
    workbook = openpyxl.load_workbook(PATH)
    worksheet = workbook.active

    data = []
    for row in range(2, worksheet.max_row + 1):
        row_data = {}

        for col in range(1, worksheet.max_column + 1):
            column_name = worksheet.cell(row=1, column=col).value
            cell_value = worksheet.cell(row=row, column=col).value
            row_data[column_name] = cell_value

        data.append(row_data)

    return data
